from django.contrib import admin
from django.utils.html import format_html
from django.utils import timezone
from django.http import HttpResponse
from .models import Preinscripcion

@admin.register(Preinscripcion)
class PreinscripcionAdmin(admin.ModelAdmin):
    list_display = ('nombre_aspirante', 'grado_interes', 'edad_display', 'fecha_solicitud', 'estado_display', 'revisado_por')
    list_filter = ('estado', 'grado_interes', 'fecha_solicitud')
    search_fields = ('nombre_aspirante', 'nombre_acudiente', 'email_contacto')
    date_hierarchy = 'fecha_solicitud'
    readonly_fields = ('fecha_solicitud', 'fecha_revision')
    
    fieldsets = (
        ('Información del Aspirante', {
            'fields': ('nombre_aspirante', 'fecha_nacimiento', 'grado_interes', 'observaciones_solicitante')
        }),
        ('Información del Acudiente', {
            'fields': ('nombre_acudiente', 'telefono_contacto', 'email_contacto')
        }),
        ('Documentos Adjuntos', {
            'fields': ('doc_identidad', 'recibo_servicios', 'certificado_sisben', 'certificados_academicos')
        }),
        ('Gestión Administrativa', {
            'fields': ('estado', 'fecha_solicitud', 'fecha_revision', 'revisado_por', 'notas_internas', 'motivo_rechazo', 'acepta_politica'),
            'classes': ('collapse',)
        }),
    )
    
    actions = ['aprobar_solicitudes', 'rechazar_solicitudes', 'aplazar_solicitudes', 'convertir_en_estudiante', 'exportar_a_excel']
    
    def edad_display(self, obj):
        """Muestra la edad del aspirante"""
        return f"{obj.get_edad()} años"
    edad_display.short_description = 'Edad'
    
    def estado_display(self, obj):
        """Muestra el estado con color"""
        color = obj.get_estado_color()
        return format_html(
            '<span style="background-color: {}; color: white; padding: 5px 10px; border-radius: 3px; font-weight: bold;">{}</span>',
            color,
            obj.get_estado_display()
        )
    estado_display.short_description = 'Estado'
    estado_display.admin_order_field = 'estado'
    
    def aprobar_solicitudes(self, request, queryset):
        """Acción masiva para aprobar solicitudes"""
        count = 0
        for solicitud in queryset:
            if solicitud.marcar_como_aprobada(request.user, "Aprobado mediante acción masiva"):
                count += 1
        self.message_user(request, f'{count} solicitud(es) aprobada(s) exitosamente.')
    aprobar_solicitudes.short_description = "✓ Aprobar solicitudes seleccionadas"
    
    def rechazar_solicitudes(self, request, queryset):
        """Acción masiva para rechazar solicitudes"""
        count = 0
        for solicitud in queryset:
            if solicitud.marcar_como_rechazada(request.user, "Rechazado mediante acción masiva"):
                count += 1
        self.message_user(request, f'{count} solicitud(es) rechazada(s).')
    rechazar_solicitudes.short_description = "✗ Rechazar solicitudes seleccionadas"
    
    def aplazar_solicitudes(self, request, queryset):
        """Acción masiva para aplazar solicitudes"""
        count = 0
        for solicitud in queryset:
            if solicitud.marcar_como_aplazada(request.user, "Aplazado mediante acción masiva"):
                count += 1
        self.message_user(request, f'{count} solicitud(es) aplazada(s).')
    aplazar_solicitudes.short_description = "⏸ Aplazar solicitudes seleccionadas"
    
    def exportar_a_excel(self, request, queryset):
        """Exporta las solicitudes seleccionadas a Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            self.message_user(request, 'Error: openpyxl no está instalado. Ejecuta: pip install openpyxl', level='error')
            return
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Solicitudes de Admisión"
        
        # Headers con estilo
        headers = ['Nombre Aspirante', 'Edad', 'Grado', 'Acudiente', 'Teléfono', 'Email', 'Estado', 'Fecha Solicitud', 'Revisado Por', 'Notas']
        ws.append(headers)
        
        # Estilo para headers
        header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for solicitud in queryset:
            ws.append([
                solicitud.nombre_aspirante,
                solicitud.get_edad(),
                solicitud.grado_interes,
                solicitud.nombre_acudiente,
                solicitud.telefono_contacto,
                solicitud.email_contacto,
                solicitud.get_estado_display(),
                solicitud.fecha_solicitud.strftime('%Y-%m-%d %H:%M'),
                solicitud.revisado_por.get_full_name() if solicitud.revisado_por else 'N/A',
                solicitud.notas_internas or ''
            ])
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Respuesta HTTP
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=solicitudes_admision_{timezone.now().strftime("%Y%m%d")}.xlsx'
        wb.save(response)
        return response
    
    exportar_a_excel.short_description = "📥 Exportar a Excel"

    def convertir_en_estudiante(self, request, queryset):
        from usuarios.models import Usuario
        from estudiantes.models import Estudiante

        count = 0
        for solicitud in queryset:
            if solicitud.estado != 'aprobado':
                continue

            username = solicitud.email_contacto.split('@')[0]
            if Usuario.objects.filter(username=username).exists():
                username = f'{username}_{solicitud.id}'

            password = Usuario.objects.make_random_password()
            user = Usuario.objects.create_user(
                username=username,
                email=solicitud.email_contacto,
                password=password,
                first_name=solicitud.nombre_aspirante.split(' ')[0],
                last_name=' '.join(solicitud.nombre_aspirante.split(' ')[1:]),
                rol='estudiante',
            )

            Estudiante.objects.create(
                usuario=user,
                fecha_nacimiento=solicitud.fecha_nacimiento,
                acudiente=solicitud.nombre_acudiente,
            )
            count += 1

        self.message_user(request, f'{count} nuevo(s) estudiante(s) creado(s).')

    convertir_en_estudiante.short_description = 'Convertir aprobados en Estudiantes'

    def save_model(self, request, obj, form, change):
        """Guarda el modelo y registra quién hizo la revisión si cambió el estado"""
        if change and 'estado' in form.changed_data and obj.estado != 'pendiente':
            if not obj.fecha_revision:
                obj.fecha_revision = timezone.now()
            if not obj.revisado_por:
                obj.revisado_por = request.user
        super().save_model(request, obj, form, change)
