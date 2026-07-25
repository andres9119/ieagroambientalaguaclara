import re
from datetime import datetime
from django.core.management.base import BaseCommand, CommandError
from django.contrib.auth import get_user_model
from django.db import transaction
from academico.models import Curso, Sede
from estudiantes.models import Estudiante, Matricula

User = get_user_model()

TIPODOC_MAP = {
    'TI': 'TI', 'TI:TARJETA DE IDENTIDAD': 'TI',
    'CC': 'CC', 'CC:CEDULA DE CIUDADANIA': 'CC',
    'CE': 'CE', 'CE:CEDULA DE EXTRANJERIA': 'CE',
    'PA': 'PA', 'PA:PASAPORTE': 'PA',
}

GENERO_MAP = {
    'MASCULINO': 'M', 'M': 'M',
    'FEMENINO': 'F', 'F': 'F',
}

ESTRATO_MAP = {
    'ESTRATO 0': '0', 'ESTRATO 1': '1', 'ESTRATO 2': '2',
    'ESTRATO 3': '3', 'ESTRATO 4': '4', 'ESTRATO 5': '5', 'ESTRATO 6': '6',
    'NO APLICA': '',
}


def parse_fecha(val):
    if not val or str(val).strip() in ('', 'NO APLICA'):
        return None
    val = str(val).strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d/%m/%Y %H:%M', '%d/%m/%Y %H:%M:%S'):
        try:
            return datetime.strptime(val, fmt).date()
        except ValueError:
            continue
    return None


def cell(col, row, key, default=''):
    idx = col.get(key)
    if idx is None:
        return default
    val = row[idx]
    return str(val).strip() if val is not None else default


def parse_float(val):
    if not val:
        return None
    try:
        return float(str(val).replace(',', '.'))
    except (ValueError, TypeError):
        return None


class Command(BaseCommand):
    help = 'Importa estudiantes desde un archivo Excel (formato SIMAT)'

    def add_arguments(self, parser):
        parser.add_argument('archivo', type=str, help='Ruta al archivo .xlsx')
        parser.add_argument('--anio', type=int, default=2026, help='Año lectivo (default: 2026)')
        parser.add_argument('--dry-run', action='store_true', help='Solo mostrar que se haría sin guardar')

    def handle(self, *args, **options):
        archivo = options['archivo']
        anio = options['anio']
        dry_run = options['dry_run']

        try:
            from openpyxl import load_workbook
        except ImportError:
            raise CommandError('openpyxl no está instalado. Instálalo con: pip install openpyxl')

        try:
            wb = load_workbook(archivo, read_only=True, data_only=True)
        except Exception as e:
            raise CommandError(f'Error al leer el archivo: {e}')

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise CommandError('El archivo está vacío')

        headers = [str(h).strip().upper() if h else '' for h in rows[0]]
        expected = ['DOC', 'TIPODOC', 'NOMBRE1', 'APELLIDO1', 'GRUPO', 'GRADO_COD']
        missing = [h for h in expected if h not in headers]
        if missing:
            raise CommandError(f'Columnas requeridas no encontradas: {missing}. Headers: {headers}')

        col = {name: headers.index(name) for name in expected if name in headers}
        for extra in ['NOMBRE2', 'APELLIDO2', 'GENERO', 'FECHA_NACIMIENTO', 'BARRIO',
                       'EPS', 'TIPO DE SANGRE', 'ESTRATO', 'NUI', 'SEDE',
                       'JORNADA', 'ZONA_SEDE', 'ETNIA', 'DISCAPACIDAD',
                       'MODELO', 'FUENTE_RECURSOS', 'CAMPESINO',
                       'CATEGORIA_AULA', 'PAIS_ORIGEN', 'ESTADO',
                       'CORREO', 'TELEFONO', 'GRADO_COD']:
            if extra in headers:
                col[extra] = headers.index(extra)

        creados = 0
        actualizados = 0
        errores = 0
        saltados = 0

        for i, row in enumerate(rows[1:], start=2):
            try:
                documento = str(row[col['DOC']]).strip()
                if not documento or documento == 'None':
                    saltados += 1
                    continue

                nombre1 = str(row[col['NOMBRE1']] or '').strip().title()
                nombre2 = cell(col, row, 'NOMBRE2').title()
                apellido1 = str(row[col['APELLIDO1']] or '').strip().title()
                apellido2 = cell(col, row, 'APELLIDO2').title()

                first_name = f'{nombre1} {nombre2}'.strip()
                last_name = f'{apellido1} {apellido2}'.strip()

                tipo_doc_raw = str(row[col['TIPODOC']] or '').strip().upper()
                tipo_doc = 'TI'
                for key, val in TIPODOC_MAP.items():
                    if key in tipo_doc_raw:
                        tipo_doc = val
                        break

                grupo = str(row[col['GRUPO']] if row[col['GRUPO']] is not None else '').strip()
                if not grupo:
                    grado_cod = cell(col, row, 'GRADO_COD')
                    grupo = grado_cod if grado_cod else 'SIN_GRUPO'

                nivel = 'BACHILLERATO'
                digits = ''.join(c for c in grupo if c.isdigit())
                if len(digits) >= 2 and digits[:2] in ('10', '11'):
                    grado = int(digits[:2])
                elif digits:
                    grado = int(digits[0])
                else:
                    grado = 0
                if 1 <= grado <= 5:
                    nivel = 'PRIMARIA'
                elif 6 <= grado <= 9:
                    nivel = 'SECUNDARIA'
                elif grado >= 10:
                    nivel = 'MEDIA'

                sede_nombre = cell(col, row, 'SEDE')

                if not sede_nombre:
                    sede_nombre = 'Sede Principal'

                if dry_run:
                    self.stdout.write(f'  [{i}] {documento} - {first_name} {last_name} -> {grupo} ({sede_nombre})')
                    continue

                with transaction.atomic():
                    sede_obj, _ = Sede.objects.get_or_create(
                        nombre__iexact=sede_nombre,
                        defaults={'nombre': sede_nombre}
                    )

                    curso = Curso.objects.filter(nombre=grupo).first()
                    if not curso:
                        curso = Curso.objects.create(
                            nombre=grupo,
                            sede=sede_obj,
                            nivel=nivel
                        )
                    elif curso.sede != sede_obj:
                        curso.sede = sede_obj
                        curso.save()

                    default_pass = documento + '*'
                    user, user_created = User.objects.get_or_create(
                        username=documento,
                        defaults={
                            'first_name': first_name,
                            'last_name': last_name,
                            'rol': 'estudiante',
                            'must_change_password': True,
                        }
                    )
                    if user_created:
                        user.set_password(default_pass)
                        user.save()
                    else:
                        user.first_name = first_name
                        user.last_name = last_name
                        if not user.has_usable_password():
                            user.set_password(default_pass)
                        user.save()
                        actualizados += 1

                    fecha_nac = parse_fecha(cell(col, row, 'FECHA_NACIMIENTO') or None)
                    genero_raw = cell(col, row, 'GENERO').upper()
                    genero = GENERO_MAP.get(genero_raw, '')
                    barrio = cell(col, row, 'BARRIO')
                    eps = cell(col, row, 'EPS')
                    rh_raw = cell(col, row, 'TIPO DE SANGRE').upper()
                    rh = rh_raw if rh_raw in dict(Estudiante.RH_CHOICES) else ''
                    estrato_raw = cell(col, row, 'ESTRATO').upper()
                    estrato = ESTRATO_MAP.get(estrato_raw, '')
                    nui = cell(col, row, 'NUI')
                    jornada = cell(col, row, 'JORNADA')
                    zona = cell(col, row, 'ZONA_SEDE')
                    etnia = cell(col, row, 'ETNIA')
                    discapacidad = cell(col, row, 'DISCAPACIDAD')
                    modelo_educativo = cell(col, row, 'MODELO')
                    fuente_recursos = cell(col, row, 'FUENTE_RECURSOS')
                    campesino_raw = cell(col, row, 'CAMPESINO').upper()
                    campesino = campesino_raw in ('S', 'SI', '1', 'TRUE')
                    categoria_aula = cell(col, row, 'CATEGORIA_AULA')
                    pais_origen = cell(col, row, 'PAIS_ORIGEN')
                    estado_matricula = cell(col, row, 'ESTADO')
                    correo = cell(col, row, 'CORREO')
                    telefono = cell(col, row, 'TELEFONO')

                    if correo and not user.email:
                        user.email = correo
                        user.save()

                    estudiante, est_created = Estudiante.objects.get_or_create(
                        usuario=user,
                        defaults={
                            'fecha_nacimiento': fecha_nac or datetime(anio, 1, 1).date(),
                            'acudiente': last_name or 'SIN ACUDIENTE',
                            'tipo_documento': tipo_doc,
                            'genero': genero,
                            'barrio': barrio,
                            'eps': eps,
                            'rh': rh,
                            'estrato': estrato,
                            'nui': nui,
                            'jornada': jornada,
                            'zona': zona,
                            'etnia': etnia,
                            'discapacidad': discapacidad,
                            'modelo_educativo': modelo_educativo,
                            'fuente_recursos': fuente_recursos,
                            'campesino': campesino,
                            'categoria_aula': categoria_aula,
                            'pais_origen': pais_origen,
                            'estado_matricula': estado_matricula,
                            'telefono': telefono,
                        }
                    )
                    if not est_created:
                        estudiante.fecha_nacimiento = fecha_nac or estudiante.fecha_nacimiento
                        estudiante.tipo_documento = tipo_doc
                        estudiante.genero = genero or estudiante.genero
                        estudiante.barrio = barrio or estudiante.barrio
                        estudiante.eps = eps or estudiante.eps
                        estudiante.rh = rh or estudiante.rh
                        estudiante.estrato = estrato or estudiante.estrato
                        estudiante.nui = nui or estudiante.nui
                        estudiante.jornada = jornada or estudiante.jornada
                        estudiante.zona = zona or estudiante.zona
                        estudiante.etnia = etnia or estudiante.etnia
                        estudiante.discapacidad = discapacidad or estudiante.discapacidad
                        estudiante.modelo_educativo = modelo_educativo or estudiante.modelo_educativo
                        estudiante.fuente_recursos = fuente_recursos or estudiante.fuente_recursos
                        estudiante.campesino = campesino
                        estudiante.categoria_aula = categoria_aula or estudiante.categoria_aula
                        estudiante.pais_origen = pais_origen or estudiante.pais_origen
                        estudiante.estado_matricula = estado_matricula or estudiante.estado_matricula
                        estudiante.telefono = telefono or estudiante.telefono
                        estudiante.save()

                    Matricula.objects.get_or_create(
                        estudiante=estudiante,
                        curso=curso,
                        anio_lectivo=anio,
                        defaults={'activo': True}
                    )

                    if est_created:
                        creados += 1

            except Exception as e:
                errores += 1
                doc_err = cell(col, row, 'DOC', '?')
                self.stderr.write(f'  Error fila {i} ({doc_err}): {e}')
                if dry_run:
                    import traceback
                    self.stderr.write(traceback.format_exc())

        total = creados + actualizados + saltados
        self.stdout.write(self.style.SUCCESS(
            f'\nResumen: {creados} creados, {actualizados} actualizados, {errores} errores, {saltados} saltados (total: {total})'
        ))
