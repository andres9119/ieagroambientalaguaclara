from datetime import time as time_type
from django.views.generic import ListView, UpdateView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Q
from django.db import transaction
from django.utils import timezone
from django.shortcuts import get_object_or_404, render, redirect
from django.contrib import messages
from django.forms import modelformset_factory
from estudiantes.models import Estudiante, Matricula
from django.views.generic.edit import CreateView
from django.urls import reverse_lazy
from .models import Curso, Materia
from .forms import CursoCreateForm, CursoEditForm, MateriaForm, MateriaEditForm
from academico.models import CursoMateria


class InfoAcademicaView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Estudiante
    template_name = 'academico/info_academica.html'
    context_object_name = 'estudiantes'
    paginate_by = 50
    
    def test_func(self):
        """Solo administradores pueden acceder"""
        return self.request.user.rol == 'admin' or self.request.user.is_superuser
    
    def get_queryset(self):
        qs = super().get_queryset().select_related('usuario').prefetch_related('calificaciones', 'asistencias')
        
        # Filtros
        curso_id = self.request.GET.get('curso')
        materia_id = self.request.GET.get('materia')
        busqueda = self.request.GET.get('q')
        
        if curso_id:
            qs = qs.filter(matriculas__curso_id=curso_id, matriculas__activo=True)
        if materia_id:
            qs = qs.filter(matriculas__curso__curso_materias__materia_id=materia_id, matriculas__activo=True)
        if busqueda:
            qs = qs.filter(
                Q(usuario__first_name__icontains=busqueda) |
                Q(usuario__last_name__icontains=busqueda) |
                Q(usuario__username__icontains=busqueda)
            )
        
        return qs.distinct()
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from academico.models import Curso, Materia
        context['cursos'] = Curso.objects.all()
        context['materias'] = Materia.objects.all()
        context['curso_seleccionado'] = self.request.GET.get('curso', '')
        context['materia_seleccionada'] = self.request.GET.get('materia', '')
        context['busqueda'] = self.request.GET.get('q', '')
        return context

@login_required
def exportar_estudiantes_excel(request):
    """Exporta la información académica a Excel"""
    if not (request.user.rol == 'admin' or request.user.rol == 'docente' or request.user.is_superuser):
        return HttpResponse("No autorizado", status=403)
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse("Error: openpyxl no está instalado", status=500)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Información Académica"
    
    # Headers
    headers = ['Nombre', 'Curso', 'Materias', 'Promedio', 'Asistencia (%)', 'Acudiente']
    ws.append(headers)
    
    # Estilo headers
    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    estudiantes = Estudiante.objects.select_related('usuario')
    
    # Aplicar filtros si existen
    curso_id = request.GET.get('curso')
    materia_id = request.GET.get('materia')
    
    if curso_id:
        estudiantes = estudiantes.filter(matriculas__curso_id=curso_id, matriculas__activo=True)
    if materia_id:
        from academico.models import CursoMateria
        cursomaterias = CursoMateria.objects.filter(materia_id=materia_id)
        estudiantes = estudiantes.filter(calificaciones__curso_materia__in=cursomaterias)
    
    for estudiante in estudiantes.distinct():
        materias_str = ", ".join([m.nombre for m in estudiante.get_materias()])
        matricula_actual = estudiante.matriculas.filter(activo=True).first()
        curso_nombre = str(matricula_actual.curso) if matricula_actual else 'Sin Curso'
        ws.append([
            estudiante.usuario.get_full_name() or estudiante.usuario.username,
            curso_nombre,
            materias_str or 'N/A',
            estudiante.get_promedio(),
            estudiante.get_porcentaje_asistencia(),
            estudiante.acudiente
        ])
    
    # Ajustar anchos
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Respuesta
    from django.utils import timezone
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=info_academica_{timezone.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    return response


class CrearCursoView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Curso
    form_class = CursoCreateForm
    template_name = 'academico/crear_curso.html'
    success_url = reverse_lazy('info_academica')

    def test_func(self):
        return self.request.user.rol == 'admin' or self.request.user.is_superuser

    def form_valid(self, form):
        # Guardar curso y luego crear relaciones CursoMateria para cada materia seleccionada
        response = super().form_valid(form)
        materias = form.cleaned_data.get('materias')
        periodo = form.cleaned_data.get('periodo_academico', '1')
        anio = form.cleaned_data.get('anio_lectivo', 2026)
        docente_universal = form.cleaned_data.get('docente_universal')
        tutor = form.cleaned_data.get('tutor')

        # Assign tutor to curso if provided
        if tutor:
            self.object.tutor = tutor
            self.object.save()

        for m in materias:
            cm, created = CursoMateria.objects.get_or_create(curso=self.object, materia=m, periodo_academico=periodo, anio_lectivo=anio)
            if docente_universal:
                cm.docente = docente_universal
                cm.save()
        return response

    def form_invalid(self, form):
        from django.contrib import messages
        messages.error(self.request, 'Corrija los errores en el formulario para crear el curso.')
        return super().form_invalid(form)


class CrearMateriaView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Materia
    form_class = MateriaForm
    template_name = 'academico/crear_materia.html'

    def test_func(self):
        return self.request.user.rol == 'admin' or self.request.user.is_superuser

    def get_success_url(self):
        from django.contrib import messages
        messages.success(self.request, f'Materia "{self.object.nombre}" creada correctamente.')
        return reverse_lazy('crear_materia')

    def form_valid(self, form):
        response = super().form_valid(form)
        docentes = form.cleaned_data.get('docentes')
        if docentes:
            self.object.docentes.set(docentes)
        return response

    def form_invalid(self, form):
        from django.contrib import messages
        messages.error(self.request, 'Corrija los errores en el formulario para crear la materia.')
        return super().form_invalid(form)


class CursoListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Curso
    template_name = 'academico/cursos_list.html'
    context_object_name = 'cursos'

    def test_func(self):
        return self.request.user.rol == 'admin' or self.request.user.is_superuser


class CursoUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Curso
    form_class = CursoEditForm
    template_name = 'academico/curso_form.html'

    def get_success_url(self):
        messages.success(self.request, f'Curso "{self.object.nombre}" actualizado correctamente.')
        return reverse_lazy('cursos_list')

    def test_func(self):
        return self.request.user.rol == 'admin' or self.request.user.is_superuser

    def form_invalid(self, form):
        from django.contrib import messages
        messages.error(self.request, 'Corrija los errores en el formulario.')
        return super().form_invalid(form)


class MateriaListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Materia
    template_name = 'academico/materias_list.html'
    context_object_name = 'materias'

    def test_func(self):
        return self.request.user.rol == 'admin' or self.request.user.is_superuser


class MateriaUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Materia
    form_class = MateriaEditForm
    template_name = 'academico/materia_form.html'

    def get_success_url(self):
        return reverse_lazy('materias_list')

    def test_func(self):
        return self.request.user.rol == 'admin' or self.request.user.is_superuser

    def form_invalid(self, form):
        from django.contrib import messages
        messages.error(self.request, 'Corrija los errores en el formulario.')
        return super().form_invalid(form)

from django.contrib.auth.decorators import login_required, user_passes_test

@login_required
@user_passes_test(lambda u: u.is_authenticated and (u.rol == 'admin' or u.is_superuser))
def eliminar_curso(request, pk):
    curso = get_object_or_404(Curso, pk=pk)
    if request.method == 'POST':
        nombre = curso.nombre
        curso.delete()
        messages.success(request, f'Curso "{nombre}" eliminado correctamente.')
        return redirect('cursos_list')
    return render(request, 'academico/confirmar_eliminar.html', {
        'object': curso,
        'titulo': 'Eliminar Curso',
        'mensaje': f'¿Estás seguro de eliminar el curso "{curso.nombre}"?',
        'cancelar_url': 'cursos_list',
    })

def is_admin(user):
    return user.is_authenticated and (user.rol == 'admin' or user.is_superuser)

@login_required
@user_passes_test(is_admin)
def asignar_docentes_curso(request, curso_id):
    """
    Vista para asignar docentes a las materias de un curso específico.
    Utiliza un formset para editar masivamente el campo 'docente' de los registros CursoMateria.
    """
    from django.shortcuts import get_object_or_404, render, redirect
    from django.contrib import messages
    from django.forms import modelformset_factory
    from academico.models import Curso, CursoMateria
    from docentes.models import Docente
    from django import forms

    curso = get_object_or_404(Curso, id=curso_id)
    
    # Asegurar que existan materias asignadas
    queryset = CursoMateria.objects.filter(curso=curso).select_related('materia', 'docente__usuario').order_by('materia__nombre')

    DocenteFormSet = modelformset_factory(
        CursoMateria,
        fields=('docente',),
        extra=0,
        widgets={
            'docente': forms.Select(attrs={'class': 'form-select form-select-sm docente-select'})
        }
    )

    if request.method == 'POST':
        formset = DocenteFormSet(request.POST, queryset=queryset)
        if formset.is_valid():
            formset.save()
            messages.success(request, f'Docentes asignados correctamente para el curso {curso.nombre}.')
            return redirect('cursos_list')
    else:
        formset = DocenteFormSet(queryset=queryset)

    docentes_totales = Docente.objects.all().select_related('usuario')

    return render(request, 'academico/asignar_docentes.html', {
        'curso': curso,
        'formset': formset,
        'docentes_totales': docentes_totales
    })


@login_required
def promocionar_curso(request, curso_id):
    if not (request.user.rol == 'admin' or request.user.is_superuser):
        return HttpResponse("No autorizado", status=403)

    curso_origen = get_object_or_404(Curso, id=curso_id)
    cursos_destino = Curso.objects.exclude(id=curso_id).order_by('nivel', 'nombre')
    prox_anio = timezone.now().year + 1

    estudiantes = Estudiante.objects.filter(
        matriculas__curso=curso_origen, matriculas__activo=True
    ).select_related('usuario').order_by('usuario__first_name', 'usuario__last_name')

    if request.method == 'POST':
        curso_destino_id = request.POST.get('curso_destino')
        anio_destino = request.POST.get('anio_destino', str(prox_anio))
        desactivar_origen = request.POST.get('desactivar_origen') == 'on'
        estudiante_ids = request.POST.getlist('estudiantes')

        if not curso_destino_id:
            messages.error(request, 'Debes seleccionar un curso de destino.')
            return redirect('promocionar_curso', curso_id=curso_id)

        if not estudiante_ids:
            messages.error(request, 'Debes seleccionar al menos un estudiante.')
            return redirect('promocionar_curso', curso_id=curso_id)

        curso_destino = get_object_or_404(Curso, id=curso_destino_id)
        count = 0

        with transaction.atomic():
            for e in estudiantes.filter(id__in=estudiante_ids):
                mat_actual = e.matriculas.filter(curso=curso_origen, activo=True).first()
                if not mat_actual:
                    continue

                if desactivar_origen:
                    mat_actual.activo = False
                    mat_actual.save()

                _, created = Matricula.objects.get_or_create(
                    estudiante=e,
                    curso=curso_destino,
                    anio_lectivo=int(anio_destino),
                    defaults={'activo': True}
                )
                if created:
                    count += 1

        messages.success(
            request,
            f'{count} estudiante(s) promocionado(s) de "{curso_origen.nombre}" a "{curso_destino.nombre}" ({anio_destino}).'
        )
        return redirect('cursos_list')

    return render(request, 'academico/promocionar_curso.html', {
        'curso_origen': curso_origen,
        'cursos_destino': cursos_destino,
        'prox_anio': prox_anio,
        'estudiantes': estudiantes,
    })


@login_required
def horarios_curso(request, curso_id):
    from .models import Horario
    DIAS = ['1', '2', '3', '4', '5']
    curso = get_object_or_404(Curso, id=curso_id)
    es_titular = request.user.rol == 'docente' and curso.tutor == getattr(request.user, 'perfil_docente', None)
    if not (request.user.is_superuser or request.user.rol == 'admin' or es_titular):
        return HttpResponse("No autorizado", status=403)

    cm_list = list(CursoMateria.objects.filter(curso=curso).select_related('materia').order_by('materia__nombre'))

    if request.method == 'POST':
        Horario.objects.filter(curso=curso).delete()
        for i in request.POST.getlist('idx'):
            inicio = request.POST.get(f'inicio_{i}')
            fin = request.POST.get(f'fin_{i}')
            if not (inicio and fin):
                continue
            for dia in DIAS:
                val = request.POST.get(f'val_{i}_{dia}')
                if not val:
                    continue
                if val in ('d1', 'd2', 'd3'):
                    Horario.objects.create(curso=curso, dia=dia,
                        hora_inicio=inicio, hora_fin=fin, es_descanso=True)
                elif val.isdigit():
                    cm = next((c for c in cm_list if c.id == int(val)), None)
                    if cm:
                        Horario.objects.create(curso=curso, curso_materia=cm,
                            dia=dia, hora_inicio=inicio, hora_fin=fin)
        messages.success(request, f'Horarios guardados para {curso.nombre}.')
        return redirect('horarios_curso', curso_id=curso_id)

    horarios_qs = list(Horario.objects.filter(curso=curso).select_related('curso_materia__materia'))

    tiempos_ordenados = sorted(set((h.hora_inicio, h.hora_fin) for h in horarios_qs), key=lambda x: x[0])

    filas = []
    for h_ini, h_fin in tiempos_ordenados:
        row = {}
        for dia in DIAS:
            h = next((x for x in horarios_qs if x.dia == dia and x.hora_inicio == h_ini and x.hora_fin == h_fin), None)
            if h and h.es_descanso:
                row[dia] = 'd1'
            elif h:
                row[dia] = str(h.curso_materia_id)
            else:
                row[dia] = ''
        filas.append(((h_ini, h_fin), row))

    descansos = []
    for i in range(curso.num_descansos):
        descansos.append({
            'num': i + 1,
            'dur': getattr(curso, f'descanso{i+1}_min', curso.duracion_descanso),
        })

    return render(request, 'academico/horarios_curso.html', {
        'curso': curso,
        'cm_list': cm_list,
        'dias': DIAS,
        'filas': filas,
        'contador': len(filas) or 1,
        'dur_clase': curso.duracion_clase,
        'descansos': descansos,
    })



